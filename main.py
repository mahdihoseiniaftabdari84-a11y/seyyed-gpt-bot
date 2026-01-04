import io
import os
import pandas as pd
import asyncpg

from aiogram.filters import Command
from aiogram.types import BufferedInputFile

import os
import re
import asyncio
import random
import string
from datetime import datetime, timedelta, timezone

import aiosqlite
import jdatetime
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook

from aiogram import Bot, Dispatcher, F
from aiogram.client.default import DefaultBotProperties
from aiogram.enums import ParseMode
from aiogram.types import (
    Message, CallbackQuery,
    ReplyKeyboardMarkup, KeyboardButton, ReplyKeyboardRemove
)
from aiogram.utils.keyboard import InlineKeyboardBuilder
from aiogram.filters import CommandStart, Command

from aiogram.fsm.state import StatesGroup, State
from aiogram.fsm.context import FSMContext


# -------------------- ENV --------------------
load_dotenv()

# ✅ مقادیر پیش‌فرضی که دادی (اگر ENV ست نباشه از اینا استفاده می‌شه)
DEFAULT_ADMIN_ID = "5303374050"
DEFAULT_BOT_TOKEN = "PUT_YOUR_BOT_TOKEN_HERE"  # ⚠️ توکن واقعی رو داخل ENV بذار
DEFAULT_CHANNEL_ID = "-1003674522523"
DEFAULT_DATABASE_URL = "postgresql://postgres:gbZOKrXWWBLWuhdyspCICBVOujEfpVwu@switchyard.proxy.rlwy.net:23439/railway"
DEFAULT_CHANNEL_LINK = "https://t.me/SEYEDGPT"

# ✅ (اضافه شد) پیش‌فرض شماره کارت و نام کارت
DEFAULT_CARD_NUMBER = "5859 8312 4336 2216"
DEFAULT_CARD_NAME = "سید مهدی حسینی "

# ✅ تنظیم صحیح از ENV (اولویت با ENV)
ADMIN_ID = int((os.getenv("ADMIN_ID", DEFAULT_ADMIN_ID) or "0").strip() or "0")
DATABASE_URL = (os.getenv("DATABASE_URL", DEFAULT_DATABASE_URL) or "").strip()

BOT_TOKEN = (os.getenv("BOT_TOKEN", DEFAULT_BOT_TOKEN) or "").strip()

CHANNEL_ID = (os.getenv("CHANNEL_ID", DEFAULT_CHANNEL_ID) or "").strip()          # مثل: -1001234567890
CHANNEL_LINK = (os.getenv("CHANNEL_LINK", DEFAULT_CHANNEL_LINK) or "").strip()   # مثل: https://t.me/YourChannel
if not CHANNEL_LINK:
    CHANNEL_LINK = DEFAULT_CHANNEL_LINK  # لینک کانال شما

# ✅ (تغییر شد) کارت‌به‌کارت: اگر ENV خالی بود، از پیش‌فرض پر شود
CARD_NUMBER = (os.getenv("CARD_NUMBER", DEFAULT_CARD_NUMBER) or "").strip()
CARD_NAME = (os.getenv("CARD_NAME", DEFAULT_CARD_NAME) or "SEYED GPT").strip()

# مسیرها
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, "db.sqlite3")
EXCEL_PATH = os.path.join(BASE_DIR, "data.xlsx")

# قیمت پلن
PLAN_TITLE = "ChatGPT Plus — 1 Month (Single User)"
PLAN_PRICE = 369_000  # تومان

TEHRAN_TZ = timezone(timedelta(hours=3, minutes=30))


async def fetch_users():
    async with aiosqlite.connect(DB_PATH) as db:
        cur = await db.execute(
            "SELECT user_id, username, full_name, created_at FROM users ORDER BY user_id DESC"
        )
        rows = await cur.fetchall()
        return [
            {
                "user_id": r[0],
                "username": r[1],
                "full_name": r[2],
                "created_at": r[3],
            }
            for r in rows
        ]

async def send_excel_to_admin(bot, rows: list[dict], filename: str = "report.xlsx"):
    df = pd.DataFrame(rows)

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="data")

    buffer.seek(0)
    file = BufferedInputFile(buffer.read(), filename=filename)

    await bot.send_document(ADMIN_ID, file, caption="📊 گزارش اکسل آماده شد.")


# -------------------- Helpers --------------------
def now_utc_iso() -> str:
    return datetime.utcnow().replace(tzinfo=timezone.utc).isoformat()

def to_jalali_str(dt: datetime) -> str:
    dt_teh = dt.astimezone(TEHRAN_TZ)
    jdt = jdatetime.datetime.fromgregorian(datetime=dt_teh.replace(tzinfo=None))
    return jdt.strftime("%Y/%m/%d %H:%M")

def safe_int(s: str, default: int = 0) -> int:
    try:
        return int(s)
    except Exception:
        return default

def is_admin(user_id: int) -> bool:
    return ADMIN_ID and user_id == ADMIN_ID

def random_discount_code(length: int = 5) -> str:
    alphabet = string.ascii_uppercase + string.digits
    return "".join(random.choice(alphabet) for _ in range(length))

def random_discount_percent() -> int:
    return random.randint(20, 40)

def calc_discounted_price(price: int, percent: int) -> int:
    return int(price * (100 - percent) / 100)

def clamp_text(s: str, max_len: int = 800) -> str:
    if s is None:
        return ""
    s = s.strip()
    s = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]", "", s)
    if len(s) > max_len:
        s = s[:max_len]
    return s

async def safe_answer(message: Message, text: str, **kwargs):
    """
    جلوگیری از ارور parse entities در MARKDOWN:
    اول با تنظیمات فعلی می‌فرستد، اگر parse مشکل داشت، دوباره بدون parse_mode می‌فرستد.
    """
    try:
        return await message.answer(text, **kwargs)
    except Exception as e:
        if "can't parse entities" in str(e):
            kwargs.pop("parse_mode", None)
            return await message.answer(text, parse_mode=None, **kwargs)
        raise

async def safe_send(bot_obj: Bot, chat_id: int, text: str, **kwargs):
    """
    safe send_message برای ادمین/کاربر
    """
    try:
        return await bot_obj.send_message(chat_id, text, **kwargs)
    except Exception as e:
        if "can't parse entities" in str(e):
            kwargs.pop("parse_mode", None)
            return await bot_obj.send_message(chat_id, text, parse_mode=None, **kwargs)
        raise


EMAIL_RE = re.compile(r"^[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}$")
PHONE_RE = re.compile(r"^(?:\+98|0)?9\d{9}$")


# -------------------- DB Migration (orders columns) --------------------
async def ensure_orders_columns(db: aiosqlite.Connection):
    needed = {
        "plan_title": "TEXT",
        "base_amount": "INTEGER",
        "discount_code": "TEXT",
        "discount_percent": "INTEGER",
        "final_amount": "INTEGER",
        "pay_method": "TEXT",
        "status": "TEXT",
        "stage": "INTEGER",
        "receipt_file_id": "TEXT",
        "approved_at": "TEXT",
        "expires_at": "TEXT",
        "email": "TEXT",
        "phone": "TEXT",
        "reward_code": "TEXT",
        "reward_percent": "INTEGER",
        "reward_issued_at": "TEXT",

        # --- ارسال اکانت GPT ---
        "gpt_username": "TEXT",
        "gpt_password": "TEXT",
        "gpt_sent_at": "TEXT",
    }

    cur = await db.execute("PRAGMA table_info(orders)")
    cols = await cur.fetchall()
    existing = {c[1] for c in cols}

    for col, coltype in needed.items():
        if col not in existing:
            await db.execute(f"ALTER TABLE orders ADD COLUMN {col} {coltype}")


# -------------------- DB Schema & Excel --------------------
async def ensure_db():
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute("""
        CREATE TABLE IF NOT EXISTS users (
            user_id INTEGER PRIMARY KEY,
            username TEXT,
            full_name TEXT,
            created_at TEXT,
            last_member_check_at TEXT
        )
        """)

        # orders حداقلی + مایگریشن
        await db.execute("""
        CREATE TABLE IF NOT EXISTS orders (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            username TEXT,
            full_name TEXT,
            created_at TEXT
        )
        """)
        await ensure_orders_columns(db)

        await db.execute("""
        CREATE TABLE IF NOT EXISTS discount_codes (
            code TEXT PRIMARY KEY,
            percent INTEGER,
            issued_to_user INTEGER,
            issued_at TEXT,
            used_by_order INTEGER,
            used_at TEXT
        )
        """)

        await db.execute("""
        CREATE TABLE IF NOT EXISTS feedback (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            username TEXT,
            full_name TEXT,
            text TEXT,
            created_at TEXT
        )
        """)

        # Support chat tables
        await db.execute("""
        CREATE TABLE IF NOT EXISTS support_threads (
            user_id INTEGER PRIMARY KEY,
            is_open INTEGER,
            opened_at TEXT,
            closed_at TEXT
        )
        """)

        await db.execute("""
        CREATE TABLE IF NOT EXISTS support_links (
            admin_message_id INTEGER PRIMARY KEY,
            user_id INTEGER,
            created_at TEXT
        )
        """)

        await db.commit()

async def upsert_user(user_id: int, username: str, full_name: str):
    async with aiosqlite.connect(DB_PATH) as db:
        cur = await db.execute("SELECT user_id FROM users WHERE user_id=?", (user_id,))
        row = await cur.fetchone()
        if row is None:
            await db.execute("""
            INSERT INTO users(user_id, username, full_name, created_at, last_member_check_at)
            VALUES (?, ?, ?, ?, ?)
            """, (user_id, username, full_name, now_utc_iso(), None))
        else:
            await db.execute("""
            UPDATE users
            SET username=?, full_name=?
            WHERE user_id=?
            """, (username, full_name, user_id))
        await db.commit()

def ensure_excel():
    if os.path.exists(EXCEL_PATH):
        return
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Orders"
    ws1.append([
        "OrderID", "UserID", "Username", "FullName",
        "Email", "Phone", "PlanTitle",
        "BaseAmount", "DiscountCode", "DiscountPercent", "FinalAmount",
        "PayMethod", "Status", "Stage",
        "CreatedAtJalali", "ApprovedAtJalali", "ExpiresAtJalali",
        "RewardCode", "RewardPercent", "RewardIssuedAtJalali",
        "GPT_Username", "GPT_Password", "GPT_SentAtJalali"
    ])
    ws2 = wb.create_sheet("Feedback")
    ws2.append(["UserID", "Username", "FullName", "Text", "CreatedAtJalali"])
    wb.save(EXCEL_PATH)

def excel_append_order(row: list):
    ensure_excel()
    try:
        wb = load_workbook(EXCEL_PATH)
        ws = wb["Orders"]
        ws.append(row)
        wb.save(EXCEL_PATH)
        wb.close()
    except PermissionError:
        fallback = EXCEL_PATH.replace(".xlsx", "_NEW.xlsx")
        print("EXCEL LOCKED. Writing to:", fallback)
        try:
            if not os.path.exists(fallback):
                wb = Workbook()
                ws = wb.active
                ws.title = "Orders"
                ws.append([
                    "OrderID", "UserID", "Username", "FullName",
                    "Email", "Phone", "PlanTitle",
                    "BaseAmount", "DiscountCode", "DiscountPercent", "FinalAmount",
                    "PayMethod", "Status", "Stage",
                    "CreatedAtJalali", "ApprovedAtJalali", "ExpiresAtJalali",
                    "RewardCode", "RewardPercent", "RewardIssuedAtJalali",
                    "GPT_Username", "GPT_Password", "GPT_SentAtJalali"
                ])
                ws.append(row)
                wb.save(fallback)
                wb.close()
            else:
                wb = load_workbook(fallback)
                ws = wb["Orders"]
                ws.append(row)
                wb.save(fallback)
                wb.close()
        except Exception as e:
            print("EXCEL FALLBACK ERROR:", e)
    except Exception as e:
        print("EXCEL WRITE ERROR:", e)

def excel_update_order(order_id: int, **updates):
    ensure_excel()
    try:
        wb = load_workbook(EXCEL_PATH)
        ws = wb["Orders"]

        target_row = None
        for r in range(2, ws.max_row + 1):
            cell_val = ws.cell(row=r, column=1).value
            if str(cell_val) == str(order_id):
                target_row = r
                break

        if not target_row:
            wb.close()
            return

        col_map = {
            "discount_code": 9,
            "discount_percent": 10,
            "final_amount": 11,
            "pay_method": 12,
            "status": 13,
            "stage": 14,
            "approved_at_jalali": 16,
            "expires_at_jalali": 17,
            "reward_code": 18,
            "reward_percent": 19,
            "reward_issued_at_jalali": 20,
            "gpt_username": 21,
            "gpt_password": 22,
            "gpt_sent_at_jalali": 23,
        }

        for k, v in updates.items():
            if k in col_map:
                ws.cell(row=target_row, column=col_map[k]).value = v

        wb.save(EXCEL_PATH)
        wb.close()
    except PermissionError:
        print("EXCEL UPDATE ERROR: file is locked (close Excel).")
    except Exception as e:
        print("EXCEL UPDATE ERROR:", e)

def excel_append_feedback(row: list):
    ensure_excel()
    try:
        wb = load_workbook(EXCEL_PATH)
        ws = wb["Feedback"]
        ws.append(row)
        wb.save(EXCEL_PATH)
        wb.close()
    except PermissionError:
        fallback = EXCEL_PATH.replace(".xlsx", "_NEW.xlsx")
        print("EXCEL LOCKED. Writing feedback to:", fallback)
        try:
            if not os.path.exists(fallback):
                wb = Workbook()
                ws1 = wb.active
                ws1.title = "Orders"
                ws1.append([
                    "OrderID", "UserID", "Username", "FullName",
                    "Email", "Phone", "PlanTitle",
                    "BaseAmount", "DiscountCode", "DiscountPercent", "FinalAmount",
                    "PayMethod", "Status", "Stage",
                    "CreatedAtJalali", "ApprovedAtJalali", "ExpiresAtJalali",
                    "RewardCode", "RewardPercent", "RewardIssuedAtJalali",
                    "GPT_Username", "GPT_Password", "GPT_SentAtJalali"
                ])
                ws2 = wb.create_sheet("Feedback")
                ws2.append(["UserID", "Username", "FullName", "Text", "CreatedAtJalali"])
                ws2.append(row)
                wb.save(fallback)
                wb.close()
            else:
                wb = load_workbook(fallback)
                if "Feedback" not in wb.sheetnames:
                    ws2 = wb.create_sheet("Feedback")
                    ws2.append(["UserID", "Username", "FullName", "Text", "CreatedAtJalali"])
                ws2 = wb["Feedback"]
                ws2.append(row)
                wb.save(fallback)
                wb.close()
        except Exception as e:
            print("EXCEL FALLBACK FEEDBACK ERROR:", e)
    except Exception as e:
        print("EXCEL WRITE ERROR:", e)


# -------------------- Channel membership check --------------------
async def is_member(bot: Bot, user_id: int) -> bool:
    if not CHANNEL_ID:
        return False
    try:
        member = await bot.get_chat_member(chat_id=int(CHANNEL_ID), user_id=user_id)
        return member.status in ("member", "administrator", "creator")
    except Exception:
        return False


# -------------------- UI --------------------
def main_menu_kb() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="🛒 خرید اشتراک"), KeyboardButton(text="💎 پلن و قیمت")],
            [KeyboardButton(text="🎟 کد تخفیف"), KeyboardButton(text="💬 نظر / پیشنهاد")],
            [KeyboardButton(text="🆘 پشتیبانی (چت)"), KeyboardButton(text="❓ سوالات متداول")]
        ],
        resize_keyboard=True
    )

# ✅ منوی مخصوص هر کاربر (فقط برای ادمین: دکمه گزارش اکسل نمایش داده می‌شود)
def main_menu_kb_for(user_id: int) -> ReplyKeyboardMarkup:
    keyboard = [
        [KeyboardButton(text="🛒 خرید اشتراک"), KeyboardButton(text="💎 پلن و قیمت")],
        [KeyboardButton(text="🎟 کد تخفیف"), KeyboardButton(text="💬 نظر / پیشنهاد")],
        [KeyboardButton(text="🆘 پشتیبانی (چت)"), KeyboardButton(text="❓ سوالات متداول")]
    ]
    if is_admin(user_id):
        keyboard.insert(0, [KeyboardButton(text="📊 گزارش اکسل")])

    return ReplyKeyboardMarkup(
        keyboard=keyboard,
        resize_keyboard=True
    )

def cancel_only_kb() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        keyboard=[[KeyboardButton(text="❌ لغو عملیات")]],
        resize_keyboard=True,
        one_time_keyboard=True
    )

def discount_step_kb() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="⏭ بدون کد تخفیف")],
            [KeyboardButton(text="❌ لغو عملیات")]
        ],
        resize_keyboard=True
    )

def payment_methods_kb() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="💳 کارت‌به‌کارت"), KeyboardButton(text="🟦 پرداخت آنلاین (به‌زودی)")],
            [KeyboardButton(text="❌ لغو عملیات")]
        ],
        resize_keyboard=True
    )

def join_channel_inline_kb() -> InlineKeyboardBuilder:
    kb = InlineKeyboardBuilder()
    kb.button(text="✅ عضویت در کانال", url=CHANNEL_LINK)
    kb.button(text="🔄 بررسی عضویت", callback_data="check_join")
    kb.adjust(1)
    return kb

def stage_text(stage: int) -> str:
    return {
        1: "1️⃣ دریافت اطلاعات",
        2: "2️⃣ آماده‌سازی",
        3: "3️⃣ ارسال اطلاعات"
    }.get(stage, "—")

def order_status_text(status: str) -> str:
    m = {
        "NEW": "🆕 جدید",
        "WAITING_PAYMENT": "💳 در انتظار پرداخت",
        "WAITING_ADMIN": "⏳ در انتظار تایید ادمین",
        "APPROVED": "✅ تایید شد",
        "REJECTED": "❌ رد شد",
        "CANCELLED": "🚫 لغو شد"
    }
    return m.get(status, status)

def admin_order_kb(order_id: int) -> InlineKeyboardBuilder:
    kb = InlineKeyboardBuilder()
    kb.button(text="✅ تایید پرداخت", callback_data=f"admin:approve:{order_id}")
    kb.button(text="❌ رد پرداخت", callback_data=f"admin:reject:{order_id}")
    kb.button(text="🚚 مرحله 1", callback_data=f"admin:stage:{order_id}:1")
    kb.button(text="🛠 مرحله 2", callback_data=f"admin:stage:{order_id}:2")
    kb.button(text="📤 مرحله 3", callback_data=f"admin:stage:{order_id}:3")
    kb.button(text="📩 ارسال اکانت GPT", callback_data=f"admin:sendacc:{order_id}")
    kb.adjust(2, 3, 1)
    return kb


# -------------------- FSM --------------------
class Flow(StatesGroup):
    waiting_email = State()
    waiting_phone = State()
    waiting_discount = State()
    waiting_payment_choice = State()
    waiting_receipt = State()
    waiting_feedback = State()
    waiting_support = State()

class AdminFlow(StatesGroup):
    waiting_gpt_credentials = State()


# -------------------- Bot init --------------------
if not BOT_TOKEN or BOT_TOKEN == "PUT_YOUR_BOT_TOKEN_HERE":
    raise RuntimeError("BOT_TOKEN is missing. Put it in .env or ENV variables.")

bot = Bot(
    token=BOT_TOKEN,
    default=DefaultBotProperties(parse_mode=ParseMode.MARKDOWN)
)
dp = Dispatcher()


# -------------------- Access control (Channel gate) --------------------
async def require_access(msg_or_cb, user_id: int) -> bool:
    if is_admin(user_id):
        return True

    if not CHANNEL_ID:
        text = (
            "⚠️ برای استفاده از ربات، ابتدا باید عضو کانال شوید.\n\n"
            f"لینک کانال:\n{CHANNEL_LINK}\n\n"
            "بعد از عضویت، روی «🔄 بررسی عضویت» بزن ✅"
        )
        if isinstance(msg_or_cb, Message):
            await safe_answer(msg_or_cb, text, reply_markup=join_channel_inline_kb().as_markup())
        else:
            await safe_answer(msg_or_cb.message, text, reply_markup=join_channel_inline_kb().as_markup())
            await msg_or_cb.answer()
        return False

    ok = await is_member(bot, user_id)
    if not ok:
        text = (
            "⚠️ برای استفاده از ربات، ابتدا باید عضو کانال شوید.\n\n"
            "بعد از عضویت، روی «🔄 بررسی عضویت» بزن ✅"
        )
        if isinstance(msg_or_cb, Message):
            await safe_answer(msg_or_cb, text, reply_markup=join_channel_inline_kb().as_markup())
        else:
            await safe_answer(msg_or_cb.message, text, reply_markup=join_channel_inline_kb().as_markup())
            await msg_or_cb.answer()
        return False

    return True


# -------------------- Support chat helpers --------------------
async def open_support(user_id: int):
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute("""
        INSERT INTO support_threads(user_id, is_open, opened_at, closed_at)
        VALUES(?, 1, ?, NULL)
        ON CONFLICT(user_id) DO UPDATE SET is_open=1, opened_at=?, closed_at=NULL
        """, (user_id, now_utc_iso(), now_utc_iso()))
        await db.commit()

async def close_support(user_id: int):
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute(
            "UPDATE support_threads SET is_open=0, closed_at=? WHERE user_id=?",
            (now_utc_iso(), user_id)
        )
        await db.commit()

async def link_admin_message(admin_message_id: int, user_id: int):
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute(
            "INSERT OR REPLACE INTO support_links(admin_message_id, user_id, created_at) VALUES(?,?,?)",
            (admin_message_id, user_id, now_utc_iso())
        )
        await db.commit()

async def get_user_by_admin_message(admin_message_id: int) -> int | None:
    async with aiosqlite.connect(DB_PATH) as db:
        cur = await db.execute("SELECT user_id FROM support_links WHERE admin_message_id=?", (admin_message_id,))
        row = await cur.fetchone()
        return int(row[0]) if row else None


# -------------------- Start / Home --------------------
@dp.message(CommandStart())
async def cmd_start(msg: Message, state: FSMContext):
    await state.clear()

    username = msg.from_user.username or ""
    full_name = (msg.from_user.full_name or msg.from_user.first_name or "").strip()
    await upsert_user(msg.from_user.id, username, full_name)

    if not await require_access(msg, msg.from_user.id):
        return

    await safe_answer(
        msg,
        "🌟 به *SEYED GPT* خوش اومدی!\n\n"
        "از منوی پایین انتخاب کن 👇",
        reply_markup=main_menu_kb_for(msg.from_user.id)
    )

@dp.message(Command("excel"))
async def cmd_excel(message):
    if message.from_user.id != ADMIN_ID:
        return await message.answer("⛔ فقط ادمین اجازه دارد.")

    rows = await fetch_users()
    await send_excel_to_admin(message.bot, rows, filename="users.xlsx")
    await message.answer("✅ فایل اکسل ارسال شد.")


@dp.callback_query(F.data == "check_join")
async def check_join(cb: CallbackQuery):
    if not CHANNEL_ID:
        await safe_answer(
            cb.message,
            "✅ ثبت شد.\n"
            "⚠️ برای چک واقعی عضویت باید CHANNEL_ID را در .env تنظیم کنی.\n"
            "فعلاً ادامه بده.",
            reply_markup=main_menu_kb_for(cb.from_user.id)
        )
        await cb.answer()
        return

    if not await require_access(cb, cb.from_user.id):
        return

    await safe_answer(cb.message, "✅ عضویت تایید شد. حالا می‌تونی از ربات استفاده کنی.", reply_markup=main_menu_kb_for(cb.from_user.id))
    await cb.answer()


# -------------------- Menu actions --------------------
@dp.message(F.text == "📊 گزارش اکسل")
async def excel_button(msg: Message):
    if not is_admin(msg.from_user.id):
        return await safe_answer(msg, "⛔ فقط ادمین اجازه دارد.", reply_markup=main_menu_kb_for(msg.from_user.id))

    if not DATABASE_URL:
        return await safe_answer(msg, "❌ DATABASE_URL تنظیم نشده.", reply_markup=main_menu_kb_for(msg.from_user.id))

    try:
        rows = await fetch_users()
        await send_excel_to_admin(msg.bot, rows, filename="users.xlsx")
        await safe_answer(msg, "✅ فایل اکسل ارسال شد.", reply_markup=main_menu_kb_for(msg.from_user.id))
    except Exception as e:
        await safe_answer(
            msg,
            f"❌ خطا در گزارش اکسل:\n{e}",
            parse_mode=None,
            reply_markup=main_menu_kb_for(msg.from_user.id)
        )

@dp.message(F.text == "💎 پلن و قیمت")
async def plans(msg: Message):
    if not await require_access(msg, msg.from_user.id):
        return
    text = (
        "💎 *پلن‌ها و قیمت‌ها*\n\n"
        f"• *{PLAN_TITLE}*\n"
        f"• قیمت: *{PLAN_PRICE:,} تومان*\n\n"
        "برای خرید روی «🛒 خرید اشتراک» بزن."
    )
    await safe_answer(msg, text, reply_markup=main_menu_kb_for(msg.from_user.id))

@dp.message(F.text == "❓ سوالات متداول")
async def faq(msg: Message):
    if not await require_access(msg, msg.from_user.id):
        return
    text = (
        "❓ *سوالات متداول*\n\n"
        "• فعال‌سازی چقدر طول می‌کشه؟\n"
        "  بعد از تایید پرداخت و پیشرفت مراحل، بهت اطلاع داده می‌شه.\n\n"
        "• روی موبایل هم کار می‌کنه؟\n"
        "  بله ✅\n\n"
        "• اگر مشکلی پیش بیاد چی؟\n"
        "  پشتیبانی داریم ✅"
    )
    await safe_answer(msg, text, reply_markup=main_menu_kb_for(msg.from_user.id))

@dp.message(F.text == "🎟 کد تخفیف")
async def discount_info(msg: Message):
    if not await require_access(msg, msg.from_user.id):
        return
    await safe_answer(
        msg,
        "🎟 کد تخفیف را *در مرحله خرید* وارد می‌کنی و همانجا از مبلغ کم می‌شود.\n"
        "برای خرید از «🛒 خرید اشتراک» استفاده کن ✅",
        reply_markup=main_menu_kb_for(msg.from_user.id)
    )

@dp.message(F.text == "💬 نظر / پیشنهاد")
async def feedback_start(msg: Message, state: FSMContext):
    if not await require_access(msg, msg.from_user.id):
        return
    await state.set_state(Flow.waiting_feedback)
    await safe_answer(
        msg,
        "💬 پیام نظر/پیشنهادت رو همینجا بفرست.\n"
        "بعد از ارسال، ثبت می‌شه ✅",
        reply_markup=cancel_only_kb()
    )

@dp.message(Flow.waiting_feedback, F.text)
async def feedback_save(msg: Message, state: FSMContext):
    text = clamp_text(msg.text or "", 1200)
    if text == "❌ لغو عملیات":
        await state.clear()
        await safe_answer(msg, "لغو شد ✅", reply_markup=main_menu_kb_for(msg.from_user.id))
        return

    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute(
            "INSERT INTO feedback(user_id, username, full_name, text, created_at) VALUES(?,?,?,?,?)",
            (msg.from_user.id, msg.from_user.username or "", msg.from_user.full_name or "", text, now_utc_iso())
        )
        await db.commit()

    excel_append_feedback([
        msg.from_user.id,
        msg.from_user.username or "",
        msg.from_user.full_name or "",
        text,
        to_jalali_str(datetime.now(tz=TEHRAN_TZ))
    ])

    await safe_answer(msg, "✅ پیام شما ثبت شد و بررسی می‌شه.", reply_markup=main_menu_kb_for(msg.from_user.id))
    await state.clear()

@dp.message(F.text == "🆘 پشتیبانی (چت)")
async def support_start(msg: Message, state: FSMContext):
    if not await require_access(msg, msg.from_user.id):
        return
    await open_support(msg.from_user.id)
    await state.set_state(Flow.waiting_support)
    await safe_answer(
        msg,
        "🆘 *پشتیبانی فعال شد*\n\n"
        "پیامت رو همینجا بفرست تا به ادمین ارسال کنم.\n"
        "برای بستن چت: «❌ لغو عملیات» را بزن.",
        reply_markup=cancel_only_kb()
    )

@dp.message(Flow.waiting_support)
async def support_user_message(msg: Message, state: FSMContext):
    if not await require_access(msg, msg.from_user.id):
        return

    if (msg.text or "").strip() == "❌ لغو عملیات":
        await close_support(msg.from_user.id)
        await state.clear()
        await safe_answer(msg, "✅ چت پشتیبانی بسته شد.", reply_markup=main_menu_kb_for(msg.from_user.id))
        return

    if not ADMIN_ID:
        await safe_answer(msg, "❌ ادمین تنظیم نشده.")
        return

    try:
        header = (
            "🆘 پیام پشتیبانی جدید\n"
            f"User: {msg.from_user.id} (@{msg.from_user.username or '-'})\n"
            f"Name: {msg.from_user.full_name or '-'}\n\n"
            "— پیام کاربر —"
        )
        await safe_send(bot, ADMIN_ID, header, parse_mode=None)

        sent = await msg.copy_to(ADMIN_ID)
        await link_admin_message(sent.message_id, msg.from_user.id)

        await safe_answer(msg, "✅ پیامت ارسال شد. منتظر پاسخ ادمین باش.", reply_markup=main_menu_kb_for(msg.from_user.id))
        await state.clear()
    except Exception as e:
        await safe_answer(msg, f"❌ ارسال به ادمین ناموفق بود: {e}", parse_mode=None)


# -------------------- Order helpers --------------------
async def create_order(user_id: int, username: str, full_name: str) -> int:
    async with aiosqlite.connect(DB_PATH) as db:
        cur = await db.execute("""
            INSERT INTO orders(
                user_id, username, full_name,
                plan_title, base_amount, discount_code, discount_percent, final_amount,
                pay_method, status, stage, receipt_file_id,
                created_at, approved_at, expires_at, email, phone,
                reward_code, reward_percent, reward_issued_at,
                gpt_username, gpt_password, gpt_sent_at
            )
            VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            user_id, username, full_name,
            PLAN_TITLE, PLAN_PRICE, None, None, PLAN_PRICE,
            None, "NEW", 1, None,
            now_utc_iso(), None, None, None, None,
            None, None, None,
            None, None, None
        ))
        await db.commit()
        return cur.lastrowid

async def update_order(order_id: int, **fields):
    if not fields:
        return
    keys = list(fields.keys())
    vals = [fields[k] for k in keys]
    sets = ", ".join([f"{k}=?" for k in keys])
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute(f"UPDATE orders SET {sets} WHERE id=?", (*vals, order_id))
        await db.commit()

async def reserve_discount(code: str) -> tuple[bool, str, int | None]:
    code = clamp_text(code, 20).upper()
    async with aiosqlite.connect(DB_PATH) as db:
        cur = await db.execute("SELECT code, percent, used_by_order FROM discount_codes WHERE code=?", (code,))
        row = await cur.fetchone()
        if not row:
            return False, "کد تخفیف معتبر نیست ❌", None
        _, percent, used_by_order = row
        if used_by_order is not None:
            return False, "این کد قبلاً استفاده شده ❌", None
        return True, "کد تخفیف تایید شد ✅", int(percent)

async def mark_discount_used(code: str, order_id: int):
    code = clamp_text(code or "", 20).upper()
    if not code:
        return
    async with aiosqlite.connect(DB_PATH) as db:
        await db.execute(
            "UPDATE discount_codes SET used_by_order=?, used_at=? WHERE code=? AND used_by_order IS NULL",
            (order_id, now_utc_iso(), code)
        )
        await db.commit()

async def issue_discount_to_user(user_id: int) -> tuple[str, int]:
    async with aiosqlite.connect(DB_PATH) as db:
        for _ in range(30):
            code = random_discount_code(5)
            percent = random_discount_percent()
            try:
                await db.execute(
                    "INSERT INTO discount_codes(code, percent, issued_to_user, issued_at, used_by_order, used_at) VALUES(?,?,?,?,?,?)",
                    (code, percent, user_id, now_utc_iso(), None, None)
                )
                await db.commit()
                return code, percent
            except Exception:
                continue
    return "SEYED1", 20


# -------------------- Buy flow --------------------
@dp.message(F.text == "🛒 خرید اشتراک")
async def buy_start(msg: Message, state: FSMContext):
    if not await require_access(msg, msg.from_user.id):
        return
    await state.clear()
    await state.set_state(Flow.waiting_email)
    await safe_answer(
        msg,
        "📧 *مرحله 1 از 2 — دریافت اطلاعات*\n"
        "لطفاً *ایمیل* را ارسال کنید.\n"
        "_مثال: name@gmail.com_",
        reply_markup=cancel_only_kb()
    )

@dp.message(Flow.waiting_email, F.text)
async def step_email(msg: Message, state: FSMContext):
    t = clamp_text(msg.text or "", 200)
    if t == "❌ لغو عملیات":
        await state.clear()
        await safe_answer(msg, "لغو شد ✅", reply_markup=main_menu_kb_for(msg.from_user.id))
        return

    if not EMAIL_RE.match(t):
        await safe_answer(msg, "❌ ایمیل معتبر نیست. مثال: name@gmail.com", reply_markup=cancel_only_kb())
        return

    await state.update_data(email=t)
    await state.set_state(Flow.waiting_phone)

    await safe_answer(
        msg,
        "📱 *مرحله 2 از 2 — دریافت اطلاعات*\n"
        "لطفاً *شماره تماس* را ارسال کنید.\n"
        "_مثال: 09123456789_",
        reply_markup=cancel_only_kb()
    )

@dp.message(Flow.waiting_phone, F.text)
async def step_phone(msg: Message, state: FSMContext):
    t = clamp_text(msg.text or "", 50)

    if t == "❌ لغو عملیات":
        data = await state.get_data()
        order_id = data.get("order_id")
        if order_id:
            await update_order(order_id, status="CANCELLED")
            excel_update_order(order_id, status="CANCELLED")
        await state.clear()
        await safe_answer(msg, "لغو شد ✅", reply_markup=main_menu_kb_for(msg.from_user.id))
        return

    if not PHONE_RE.match(t):
        await safe_answer(msg, "❌ شماره تماس معتبر نیست. مثال: 09123456789", reply_markup=cancel_only_kb())
        return

    data = await state.get_data()
    email = data.get("email")

    order_id = await create_order(
        user_id=msg.from_user.id,
        username=msg.from_user.username or "",
        full_name=msg.from_user.full_name or ""
    )
    await update_order(order_id, email=email, phone=t, status="WAITING_PAYMENT", stage=1)
    await state.update_data(order_id=order_id)

    # ✅ ثبت سفارش همان لحظه در اکسل + پایان تقریبی 30 روز از زمان ثبت
    try:
        created_dt = datetime.now(tz=timezone.utc)
        expires_dt = created_dt + timedelta(days=30)
        excel_append_order([
            order_id, msg.from_user.id, msg.from_user.username or "", msg.from_user.full_name or "",
            email or "", t, PLAN_TITLE,
            PLAN_PRICE, "", 0, PLAN_PRICE,
            "", "WAITING_PAYMENT", 1,
            to_jalali_str(created_dt),
            "", to_jalali_str(expires_dt),
            "", 0, "",
            "", "", ""
        ])
    except Exception as e:
        print("EXCEL ORDER APPEND ERROR:", e)

    await state.set_state(Flow.waiting_discount)
    await safe_answer(
        msg,
        f"✅ اطلاعات ثبت شد.\n"
        f"🧾 شماره سفارش: *{order_id}*\n\n"
        "🎟 اگر *کد تخفیف* داری همینجا بفرست.\n"
        "اگر نداری از دکمه «⏭ بدون کد تخفیف» استفاده کن.",
        reply_markup=discount_step_kb()
    )

@dp.message(Flow.waiting_discount, F.text)
async def step_discount(msg: Message, state: FSMContext):
    t = clamp_text(msg.text or "", 50)

    if t == "❌ لغو عملیات":
        data = await state.get_data()
        order_id = data.get("order_id")
        if order_id:
            await update_order(order_id, status="CANCELLED")
            excel_update_order(order_id, status="CANCELLED")
        await state.clear()
        await safe_answer(msg, "لغو شد ✅", reply_markup=main_menu_kb_for(msg.from_user.id))
        return

    data = await state.get_data()
    order_id = data.get("order_id")
    if not order_id:
        await state.clear()
        await safe_answer(msg, "مشکلی پیش اومد. دوباره از خرید شروع کن.", reply_markup=main_menu_kb_for(msg.from_user.id))
        return

    if t == "⏭ بدون کد تخفیف":
        await update_order(order_id, final_amount=PLAN_PRICE)
        excel_update_order(order_id, final_amount=PLAN_PRICE)
        await safe_answer(msg, f"✅ ادامه می‌دیم.\n💰 مبلغ نهایی: *{PLAN_PRICE:,} تومان*")
    else:
        ok, m, percent = await reserve_discount(t)
        if not ok:
            await safe_answer(msg, m + "\nیا از «⏭ بدون کد تخفیف» استفاده کن.", reply_markup=discount_step_kb())
            return

        final_amount = calc_discounted_price(PLAN_PRICE, percent)
        await update_order(
            order_id,
            discount_code=t.upper(),
            discount_percent=percent,
            final_amount=final_amount
        )
        excel_update_order(
            order_id,
            discount_code=t.upper(),
            discount_percent=int(percent),
            final_amount=int(final_amount)
        )
        await safe_answer(msg, f"{m}\n💰 مبلغ بعد از تخفیف: *{final_amount:,} تومان*")

    await state.set_state(Flow.waiting_payment_choice)
    await safe_answer(msg, "💳 *روش پرداخت* را انتخاب کن:", reply_markup=payment_methods_kb())

@dp.message(Flow.waiting_payment_choice, F.text)
async def payment_choice(msg: Message, state: FSMContext):
    t = (msg.text or "").strip()

    data = await state.get_data()
    order_id = data.get("order_id")
    if not order_id:
        await state.clear()
        await safe_answer(msg, "مشکلی پیش اومد. دوباره از خرید شروع کن.", reply_markup=main_menu_kb_for(msg.from_user.id))
        return

    if t == "❌ لغو عملیات":
        await update_order(order_id, status="CANCELLED")
        excel_update_order(order_id, status="CANCELLED")
        await state.clear()
        await safe_answer(msg, "لغو شد ✅", reply_markup=main_menu_kb_for(msg.from_user.id))
        return

    if t == "🟦 پرداخت آنلاین (به‌زودی)":
        await update_order(order_id, pay_method="GATEWAY", status="WAITING_ADMIN")
        excel_update_order(order_id, pay_method="GATEWAY", status="WAITING_ADMIN")
        await state.clear()
        await safe_answer(
            msg,
            "🟦 *پرداخت آنلاین*\n\n"
            "این روش فعلاً *به‌زودی* فعال می‌شود.\n"
            "اگر قصد پرداخت دارید، گزینه «💳 کارت‌به‌کارت» را انتخاب کنید.",
            reply_markup=main_menu_kb_for(msg.from_user.id)
        )
        if ADMIN_ID:
            try:
                await safe_send(
                    bot,
                    ADMIN_ID,
                    "🟦 درخواست پرداخت آنلاین (به‌زودی)\n"
                    f"OrderID: {order_id}\n"
                    f"User: {msg.from_user.id} (@{msg.from_user.username or '-'})",
                    reply_markup=admin_order_kb(order_id).as_markup(),
                    parse_mode=None
                )
            except Exception:
                pass
        return

    if t == "💳 کارت‌به‌کارت":
        await update_order(order_id, pay_method="CARD", status="WAITING_ADMIN")
        excel_update_order(order_id, pay_method="CARD", status="WAITING_ADMIN")
        async with aiosqlite.connect(DB_PATH) as db:
            cur = await db.execute("SELECT final_amount FROM orders WHERE id=?", (order_id,))
            row = await cur.fetchone()
            final_amount = int(row[0]) if row and row[0] is not None else PLAN_PRICE

        await state.set_state(Flow.waiting_receipt)
        await safe_answer(
            msg,
            "💳 *کارت‌به‌کارت*\n\n"
            f"🧾 سفارش: *{order_id}*\n"
            f"💰 مبلغ قابل پرداخت: *{final_amount:,} تومان*\n\n"
            "✅ لطفاً مبلغ را به شماره کارت زیر واریز کن و سپس *عکس رسید* را ارسال کن:\n\n"
            f"شماره کارت:\n`{CARD_NUMBER}`\n"
            f"به نام: *{CARD_NAME}*",
            reply_markup=cancel_only_kb()
        )
        return

    await safe_answer(msg, "لطفاً یکی از گزینه‌های روش پرداخت را از پایین صفحه انتخاب کن 👇", reply_markup=payment_methods_kb())

@dp.message(Flow.waiting_receipt, F.photo)
async def receipt_photo(msg: Message, state: FSMContext):
    data = await state.get_data()
    order_id = data.get("order_id")
    if not order_id:
        await state.clear()
        await safe_answer(msg, "مشکلی پیش اومد. دوباره از خرید شروع کن.", reply_markup=main_menu_kb_for(msg.from_user.id))
        return

    file_id = msg.photo[-1].file_id
    await update_order(order_id, receipt_file_id=file_id, status="WAITING_ADMIN")
    excel_update_order(order_id, status="WAITING_ADMIN")

    await safe_answer(msg, "✅ رسید دریافت شد.\n⏳ *در انتظار تایید ادمین* ", reply_markup=main_menu_kb_for(msg.from_user.id))
    await state.clear()

    if ADMIN_ID:
        try:
            async with aiosqlite.connect(DB_PATH) as db:
                cur = await db.execute("""
                SELECT user_id, username, full_name, email, phone, final_amount,
                       discount_code, discount_percent, pay_method, stage, status
                FROM orders WHERE id=?
                """, (order_id,))
                row = await cur.fetchone()

            if row:
                user_id, username, full_name, email, phone, final_amount, dcode, dperc, pay_method, stage, status = row
                await safe_send(
                    bot,
                    ADMIN_ID,
                    "🧾 رسید جدید برای تایید\n"
                    f"OrderID: {order_id}\n"
                    f"User: {user_id} (@{username or '-'})\n"
                    f"Name: {full_name or '-'}\n"
                    f"Email: {email or '-'}\n"
                    f"Phone: {phone or '-'}\n"
                    f"Amount: {int(final_amount or 0):,} تومان\n"
                    f"Discount: {dcode or '-'} ({dperc or 0}%)\n"
                    f"Pay: {pay_method}\n"
                    f"Status: {status}\n"
                    f"Stage: {stage_text(int(stage or 1))}",
                    reply_markup=admin_order_kb(order_id).as_markup(),
                    parse_mode=None
                )
                await bot.send_photo(ADMIN_ID, file_id)
        except Exception:
            pass

@dp.message(Flow.waiting_receipt, F.text)
async def receipt_text(msg: Message, state: FSMContext):
    t = (msg.text or "").strip()
    if t == "❌ لغو عملیات":
        data = await state.get_data()
        order_id = data.get("order_id")
        if order_id:
            await update_order(order_id, status="CANCELLED")
            excel_update_order(order_id, status="CANCELLED")
        await state.clear()
        await safe_answer(msg, "لغو شد ✅", reply_markup=main_menu_kb_for(msg.from_user.id))
        return
    await safe_answer(msg, "📸 لطفاً *عکس رسید* را ارسال کن یا «❌ لغو عملیات» بزن.", reply_markup=cancel_only_kb())


# -------------------- Admin: inline buttons --------------------
@dp.callback_query(F.data.startswith("admin:"))
async def admin_buttons(cb: CallbackQuery, state: FSMContext):
    if not is_admin(cb.from_user.id):
        await cb.answer("اجازه نداری.", show_alert=True)
        return

    parts = cb.data.split(":")
    action = parts[1]
    order_id = safe_int(parts[2], 0)

    if action == "approve":
        await admin_do_approve(order_id, cb)
    elif action == "reject":
        await admin_do_reject(order_id, cb)
    elif action == "stage":
        stage = safe_int(parts[3], 1) if len(parts) >= 4 else 1
        stage = 1 if stage < 1 else 3 if stage > 3 else stage
        await admin_set_stage(order_id, stage, cb)
    elif action == "sendacc":
        await state.set_state(AdminFlow.waiting_gpt_credentials)
        await state.update_data(admin_sendacc_order_id=order_id)
        await safe_answer(
            cb.message,
            "📩 ارسال اکانت GPT\n\n"
            "لطفاً در یک پیام به این شکل بفرست:\n"
            "user | pass\n\n"
            "مثال:\n"
            "abc@gmail.com | 12345678",
            parse_mode=None
        )

    await cb.answer()


# -------------------- Admin: دریافت user/pass و ارسال برای مشتری --------------------
@dp.message(AdminFlow.waiting_gpt_credentials, F.text)
async def admin_receive_gpt_credentials(msg: Message, state: FSMContext):
    if not is_admin(msg.from_user.id):
        return

    data = await state.get_data()
    order_id = data.get("admin_sendacc_order_id")
    if not order_id:
        await state.clear()
        await safe_answer(msg, "❌ سفارش مشخص نیست. دوباره اقدام کن.", parse_mode=None)
        return

    text = clamp_text(msg.text or "", 500)
    if "|" not in text:
        await safe_answer(msg, "❌ فرمت اشتباهه. باید مثل این باشه: user | pass", parse_mode=None)
        return

    user_part, pass_part = text.split("|", 1)
    gpt_user = clamp_text(user_part, 120)
    gpt_pass = clamp_text(pass_part, 120)

    async with aiosqlite.connect(DB_PATH) as db:
        cur = await db.execute("SELECT user_id FROM orders WHERE id=?", (order_id,))
        row = await cur.fetchone()

    if not row:
        await state.clear()
        await safe_answer(msg, "❌ سفارش پیدا نشد.", parse_mode=None)
        return

    target_user_id = int(row[0])

    sent_at = now_utc_iso()
    await update_order(order_id, gpt_username=gpt_user, gpt_password=gpt_pass, gpt_sent_at=sent_at)
    excel_update_order(
        order_id,
        gpt_username=gpt_user,
        gpt_password=gpt_pass,
        gpt_sent_at_jalali=to_jalali_str(datetime.fromisoformat(sent_at))
    )

    try:
        await safe_send(
            bot,
            target_user_id,
            "اکانت GPT شما آماده شد\n\n"
            f"User: {gpt_user}\n"
            f"Password: {gpt_pass}\n\n"
            "اگر مشکلی بود از پشتیبانی پیام بده.",
            parse_mode=None
        )
    except Exception as e:
        await safe_answer(msg, f"❌ ارسال به مشتری ناموفق بود: {e}", parse_mode=None)
        await state.clear()
        return

    await safe_answer(msg, "✅ اکانت برای مشتری ارسال شد.", parse_mode=None)
    await state.clear()


# -------------------- Admin approve/reject/stage --------------------
async def admin_do_approve(order_id: int, msg_or_cb):
    async with aiosqlite.connect(DB_PATH) as db:
        cur = await db.execute("""
        SELECT id, user_id, username, full_name, email, phone,
               discount_code, discount_percent, final_amount, pay_method, stage, created_at
        FROM orders WHERE id=?
        """, (order_id,))
        row = await cur.fetchone()

    if not row:
        if isinstance(msg_or_cb, CallbackQuery):
            await safe_answer(msg_or_cb.message, "❌ سفارش پیدا نشد.", parse_mode=None)
        else:
            await safe_answer(msg_or_cb, "❌ سفارش پیدا نشد.", parse_mode=None)
        return

    oid, user_id, username, full_name, email, phone, dcode, dperc, final_amount, pay_method, stage, created_at = row

    approved_at = datetime.now(tz=timezone.utc)
    expires_at = approved_at + timedelta(days=30)

    await update_order(
        oid,
        status="APPROVED",
        approved_at=approved_at.isoformat(),
        expires_at=expires_at.isoformat()
    )

    excel_update_order(
        oid,
        status="APPROVED",
        stage=int(stage or 1),
        approved_at_jalali=to_jalali_str(approved_at),
        expires_at_jalali=to_jalali_str(expires_at),
        pay_method=(pay_method or ""),
        final_amount=int(final_amount or 0),
        discount_code=(dcode or ""),
        discount_percent=int(dperc or 0),
    )

    if dcode:
        await mark_discount_used(dcode, oid)

    try:
        await safe_send(
            bot,
            int(user_id),
            "پرداخت شما تایید شد\n\n"
            f"سفارش: {oid}\n"
            f"مبلغ: {int(final_amount or 0):,} تومان\n"
            f"مرحله فعلی: {stage_text(int(stage or 1))}\n"
            f"شروع: {to_jalali_str(approved_at)}\n"
            f"پایان: {to_jalali_str(expires_at)}",
            parse_mode=None
        )
    except Exception:
        pass

    if isinstance(msg_or_cb, CallbackQuery):
        await safe_answer(msg_or_cb.message, f"✅ تایید شد. OrderID: {oid}", parse_mode=None)
    else:
        await safe_answer(msg_or_cb, f"✅ تایید شد. OrderID: {oid}", parse_mode=None)

async def admin_do_reject(order_id: int, msg_or_cb):
    async with aiosqlite.connect(DB_PATH) as db:
        cur = await db.execute("SELECT id, user_id, final_amount FROM orders WHERE id=?", (order_id,))
        row = await cur.fetchone()

    if not row:
        if isinstance(msg_or_cb, CallbackQuery):
            await safe_answer(msg_or_cb.message, "❌ سفارش پیدا نشد.", parse_mode=None)
        else:
            await safe_answer(msg_or_cb, "❌ سفارش پیدا نشد.", parse_mode=None)
        return

    oid, user_id, final_amount = row
    await update_order(oid, status="REJECTED")
    excel_update_order(oid, status="REJECTED")

    try:
        await safe_send(
            bot,
            int(user_id),
            "پرداخت شما تایید نشد\n\n"
            f"سفارش: {oid}\n"
            f"مبلغ: {int(final_amount or 0):,} تومان\n\n"
            "اگر فکر می‌کنی اشتباه شده، از پشتیبانی پیام بده.",
            parse_mode=None
        )
    except Exception:
        pass

    if isinstance(msg_or_cb, CallbackQuery):
        await safe_answer(msg_or_cb.message, "❌ رد شد.", parse_mode=None)
    else:
        await safe_answer(msg_or_cb, "❌ رد شد.", parse_mode=None)

async def admin_set_stage(order_id: int, stage: int, msg_or_cb):
    async with aiosqlite.connect(DB_PATH) as db:
        cur = await db.execute("""
        SELECT id, user_id, status, reward_code
        FROM orders WHERE id=?
        """, (order_id,))
        row = await cur.fetchone()

    if not row:
        if isinstance(msg_or_cb, CallbackQuery):
            await safe_answer(msg_or_cb.message, "❌ سفارش پیدا نشد.", parse_mode=None)
        else:
            await safe_answer(msg_or_cb, "❌ سفارش پیدا نشد.", parse_mode=None)
        return

    oid, user_id, status, reward_code = row
    await update_order(oid, stage=stage)
    excel_update_order(oid, stage=int(stage))

    try:
        await safe_send(
            bot,
            int(user_id),
            f"وضعیت سفارش شما بروزرسانی شد\nسفارش: {oid}\nمرحله فعلی: {stage_text(stage)}",
            parse_mode=None
        )
    except Exception:
        pass

    # ✅ کد تخفیف رندومی هدیه: فقط وقتی مرحله 3 شد و سفارش تایید شده باشد و قبلاً هدیه نداده باشیم
    if int(stage) == 3 and status == "APPROVED" and not reward_code:
        new_code, new_percent = await issue_discount_to_user(int(user_id))
        issued_at_iso = datetime.now(tz=timezone.utc).isoformat()
        await update_order(
            oid,
            reward_code=new_code,
            reward_percent=new_percent,
            reward_issued_at=issued_at_iso
        )
        excel_update_order(
            oid,
            reward_code=new_code,
            reward_percent=int(new_percent),
            reward_issued_at_jalali=to_jalali_str(datetime.fromisoformat(issued_at_iso))
        )

        try:
            await safe_send(
                bot,
                int(user_id),
                "سفارش شما تکمیل شد (مرحله 3)\n\n"
                f"کد تخفیف یک‌بار مصرف شما: {new_code}\n"
                f"میزان تخفیف: {new_percent}%",
                parse_mode=None
            )
        except Exception:
            pass

    if isinstance(msg_or_cb, CallbackQuery):
        await safe_answer(msg_or_cb.message, f"✅ مرحله شد: {stage_text(stage)}", parse_mode=None)
    else:
        await safe_answer(msg_or_cb, f"✅ مرحله شد: {stage_text(stage)}", parse_mode=None)


# -------------------- Admin reply router (support relay) --------------------
@dp.message()
async def admin_reply_router(msg: Message):
    if not is_admin(msg.from_user.id):
        return
    if not msg.reply_to_message:
        return

    user_id = await get_user_by_admin_message(msg.reply_to_message.message_id)
    if not user_id:
        return

    try:
        await msg.copy_to(user_id)
        await safe_answer(msg, "✅ ارسال شد.", parse_mode=None)
    except Exception as e:
        await safe_answer(msg, f"❌ ارسال نشد: {e}", parse_mode=None)


# -------------------- Global cancel (anywhere) --------------------
@dp.message(F.text == "❌ لغو عملیات")
async def global_cancel(msg: Message, state: FSMContext):
    data = await state.get_data()
    order_id = data.get("order_id")
    if order_id:
        await update_order(order_id, status="CANCELLED")
        excel_update_order(order_id, status="CANCELLED")
    await state.clear()
    await safe_answer(msg, "لغو شد ✅", reply_markup=main_menu_kb_for(msg.from_user.id))


# -------------------- Fallback --------------------
@dp.message()
async def fallback(msg: Message):
    if not await require_access(msg, msg.from_user.id):
        return
    await safe_answer(msg, "از منوی پایین انتخاب کن 👇", reply_markup=main_menu_kb_for(msg.from_user.id))


# -------------------- Main --------------------
async def main():
    ensure_excel()
    await ensure_db()
    await dp.start_polling(bot)

if __name__ == "__main__":
    asyncio.run(main())


